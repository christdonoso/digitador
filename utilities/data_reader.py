"""
module with utility functions to import data
"""

import csv
import datetime
#from errors.data_errors import WrongExtention
import pandas as pd
from openpyxl import load_workbook


month_to_str = lambda x : x if x > 10 else f'0{x}'
day_to_str = lambda x : x if x > 10 else f'0{x}'

HEADER = ['rut', 'run', 'rbd', 'fecha evaluacion', 'peso', 'talla', 'menarquia', 'diagnostico']


def open_csv(root, delim=',', new_line='\n')-> list:
    """
    funcion para cargar los datos desde un csv, necesita la ruta del archivo
    delimitador y nueva linea son opcionales.
    """        
    with open (root, new_line, encoding='utf-8') as file:
        data = csv.reader(file, delimiter=delim)
        lista = list(data)
    return lista        


def validate_data(data:list) -> list:
    """
    funcion que toma la data , y los deja en formato que necesitan para poder
    de entero ser digitados en plafaorma, los enteros, float y las fechas quedan como strings
    """
    valid_data = []

    for row in data:

        valid_row = []
        for item in row:
            if isinstance(item, float) or isinstance(item, int):
                valid_row.append(str(int(item)))

            elif isinstance(item, str):
                valid_row.append(item)

            elif isinstance(item, datetime.date):
                valid_row.append(f'{day_to_str(item.day)}{month_to_str(item.month)}{item.year}')
            else:
                valid_row.append(item)
        
        valid_data.append(valid_row)

    return valid_data


def validate_dataframe(valid_data:list, header:list)-> pd.DataFrame:
    """
    Funcion que recibe una lista con la data valida como lista y devuelve un dataframe 
    validado
    """
    valid_header = [item.upper() for item in header if item in valid_data[0] or item.upper() in valid_data[0]]
    print(valid_header)
    df = pd.DataFrame(valid_data[1:], columns=valid_data[0])[valid_header]

    return df


def delete_nonattendants(data:pd.DataFrame)-> pd.DataFrame:
    """
    Elimina las filas que tienen sin regstro el pedo,talla,menaruia y el diagnostico
    """
    for idx, row in data.iterrows():
        if all(pd.isna(row[col]) for col in ['PESO', 'TALLA', 'MENARQUIA', 'DIAGNOSTICO']):
            data.drop(index=idx, inplace=True)

    return data


def open_excel(root:str, sheet_name:object=None):
    wb = load_workbook(root)

    sheet_name = sheet_name if sheet_name else wb.sheetnames[0]
    #sheet_name = sheet_name if sheet_name else [sheet_name for sheet_name in wb.sheetnames]
    
    sheet = wb[sheet_name]
    data = []

    # Itera sobre las filas y columnas del rango deseado en tu hoja de Excel
    for row in sheet.iter_rows(values_only=True):
        data.append(row)

    valid_data = validate_data(data)
    valid_dataframe = validate_dataframe(valid_data, HEADER)
    only_attended = delete_nonattendants(valid_dataframe)
    print('no esta devoviendo solo el head')
    return only_attended


def get_data(root, delim=',', new_line='\n'):
    """
    Funcion para cargar la data, esta puede venir en formato CSV o 
    en un XLSX, de lo contratio, da un error
    """
    if root[-3:] == 'csv':
        return open_csv(root, delim, new_line)
    
    elif root[-4:] == 'xlsx':
        return open_excel(root)

#    else:
#        raise WrongExtention(f'el archivo {root} tiene extension incorrecta')


if __name__ == '__main__':
    data = open_excel('Comuna Osorno.xlsx')
    print(data)