"""
module with utility functions
"""

import csv
from errors import WrongExtention
from openpyxl import load_workbook


def validate_row(data:list|tuple) -> list:

    HEADER = ['rut','peso', 'talla', 'menarquia', 'rbd', 'diagnostico', 
              'fecha', 'fecha evaluacion'
            ]

    file_header = data[0]
    header_index = [index for index, value in enumerate(file_header) if value.lower() in HEADER]
    return header_index


def open_csv(root, delim=',', new_line='\n')-> list:
    """_summary_
    metodo para cargar los datos desde un csv
    Args:
        root (str): ruta del archivo
        delim (str): simbolo de delimitacion del archivo
    Returns:
        _type_: _description_
    """        

    with open (root, new_line, encoding='utf-8') as file:
        data = csv.reader(file, delimiter = f'{delim}')
        lista = list(data)
    return lista        


def validate_data(data:list, valid_index:list) -> list:
    valid_data = []
    for line in data:

        valid_line = []
        for index, item in enumerate(line):
            if index in valid_index:
                if isinstance(item, float):
                    valid_line.append(int(item))
        
        valid_data.append(valid_line)


def open_excel(root:str, sheet_name:object=None):
    # Carga el archivo Excel
    wb = load_workbook(root)

    # Selecciona la hoja de Excel que quieres leer
    sheet_name = sheet_name if sheet_name else wb.sheetnames[0]
    sheet = wb[sheet_name]
    data = []

    # Itera sobre las filas y columnas del rango deseado en tu hoja de Excel
    for row in sheet.iter_rows(values_only=True):
        data.append(row)

    valid_index = validate_row(data)
    valid_data = validate_data(data, valid_index)

    return valid_data



def get_data(root, delim=',', new_line='\n'):
    if root[-3:] == 'csv':
        return open_csv(root, delim, new_line)
    
    elif root[-4:] == 'xlsx':
        return open_excel(root)

    else:
        raise WrongExtention(f'el archivo {root} tiene extension incorrecta')



if __name__ == '__main__':
    data = open_excel('Tamizaje Los Rios 2024.xlsx')

    print(validate_row(data))
    for i in data:
        print(i)